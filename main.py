#! /usr/bin/env python
# main.py - convert staff services staff rosters into a more useful form

import os
import re
import sys
import logging
import argparse
import datetime
import io
import typing
import base64
import pathlib
import requests
import requests.exceptions
import textwrap

import xlrd
import dotenv

import openpyxl
import openpyxl.utils
import openpyxl.styles
import openpyxl.styles.colors
import openpyxl.writer.excel
import openpyxl.utils.cell
#import O365.excel

import config as config_static
import neil_tools
import arc_o365
#import o365_staffing
from arc_o365 import configuration
from neil_tools import spreadsheet_tools


# index of column label row, origin zero
STAFF_ROSTER_LABEL_ROW = 5
ORIG_SHEET_NAME = "Orig"
REPORT_DATE_FORMAT = "%Y-%m-%d %H-%M-%S %Z"

SHEET_COLOR = '99ff99'


NOW = datetime.datetime.now().astimezone()
NOW_NO_TZ = datetime.datetime.now()


def main() -> None:
    args = parse_args()
    if args.debug:
        logging.getLogger().setLevel(logging.DEBUG)
    log.debug("running...")

    # read static configuration
    config = neil_tools.init_config(config_static, ".env")

    # make sure we know about this DRO
    dr_config = config.DRConfig.lookup_dr(args.dr_id)
    if dr_config is None:
        log.fatal(f"no configuration on file for '{ args.dr_id }'")
        sys.exit(1)

    o365 = arc_o365.arc_o365.arc_o365(config, token_filename=config.TOKEN_FILENAME, timezone="America/Los_Angeles")

    # set up o365_config early, so we can add retrying failed transactions
    o365_config = configuration.O365Config(dr_config, o365.account)

    report_dict = o365.fetch_workforce_reports(dr_config.dr_id, subject_match_string=dr_config.subject_match_string)
    report_date = report_dict['created']
    report_date_stamp = report_date.strftime(REPORT_DATE_FORMAT)
    log.debug(f"report date is '{ report_date }', stamp '{ report_date_stamp }'")

    errors = False
    book_out = openpyxl.Workbook()

    # do the 'orig' roster first so it is at the end of the list
    sheet_name_roster0 = 'Roster0'
    sheet_name_roster1 = 'Roster1'
    # the cumulative roster has a bug: people assigned more than once doesn't show the current assignment.
    # use the checked in roster instead
    sheet_orig = read_roster(book_out, ORIG_SHEET_NAME, report_dict['Staff Roster - Cumulative'], STAFF_ROSTER_LABEL_ROW, ROSTER_FIXUPS)
    sheet_roster0 = read_roster(book_out, sheet_name_roster0, report_dict['Staff Roster - Checked In'], STAFF_ROSTER_LABEL_ROW, ROSTER_FIXUPS)
    sheet_roster1 = copy_sheet(dr_config, book_out, sheet_roster0, STAFF_ROSTER_LABEL_ROW, sheet_name_roster1, filter_row_active, ROSTER_FIXUPS)

    # delete column 'I': the Released column
    # first delete the column
    sheet_roster1.delete_cols(openpyxl.utils.cell.column_index_from_string('I'), 1)
    # adjust the table size
    sheet_name, table_range = o365_config.find_table_by_name(sheet_name_roster1, wb=book_out)
    log.debug(f"sheet { sheet_name } range { table_range }")

    # get the range in terms of min/max
    # tuple is min_col, min_row, max_col, max_row
    tuple_range = openpyxl.utils.cell.range_boundaries(table_range)

    log.debug(f"table range_boundaries { tuple_range }")
    # convert back to A1:Z999 style format, reducing the max column by one for the deleted column
    min_col = openpyxl.utils.get_column_letter(tuple_range[0])
    min_row = tuple_range[1]
    max_col = openpyxl.utils.get_column_letter(tuple_range[2] -1)
    max_row = tuple_range[3]
    table_ref = f"{min_col}{min_row}:{max_col}{max_row}"
    log.debug(f"resizing table: table_ref '{ table_ref }'")
    table = openpyxl.worksheet.table.Table(displayName=sheet_name_roster1, ref=table_ref)
    # delete the old table
    del sheet_roster1.tables[sheet_name_roster1]
    # add the new one
    sheet_roster1.add_table(table)

    # and copy it to a new sheet to fix up all the column header widths
    sheet_roster = copy_sheet(dr_config, book_out, sheet_roster1, 0, "Roster", {}, ROSTER_FIXUPS)
    del book_out[sheet_name_roster0]
    del book_out[sheet_name_roster1]


    read_roster(book_out, 'StaffRequests', report_dict['Open Staff Requests'], 1, ROSTER_FIXUPS)
    read_roster(book_out, 'Shifts', report_dict['DRO Shift Tool - Shift Registrant Details'], 3, SHIFTS_FIXUPS)
    read_roster(book_out, 'Air', report_dict['Air Travel Roster'], 2, AIR_FIXUPS, freeze_col="C", suppress_columns={'V':True})
    read_roster(book_out, 'Arrival', report_dict['Arrival Roster'], 4, ARRIVAL_FIXUPS, suppress_columns={'Z':True})

    sps_sheets = {
        "Late_Checkin": filter_row_checkin,
        "Need_SMS": filter_row_sms,
        "Needs_Sup": filter_row_needs_sup,
        "Days_2": filter_row_2_days,
        "Outprocess": filter_row_overstayed,
    }

    all_sheets = {
        "OM": filter_row_om,
        "WF": filter_row_wf,
        "IP": filter_row_ip,
        "ER": filter_row_er,
        "LOG": filter_row_log,
        "CC": filter_row_cc,
        "MC": filter_row_mc,
        }

    # make all the SPS sheets
    for sheet_name, filter in sps_sheets.items():
        copy_sheet(dr_config, book_out, sheet_roster, 0, sheet_name, filter, ROSTER_FIXUPS, sheet_color=SHEET_COLOR)
    for sheet_name, filter in all_sheets.items():
        copy_sheet(dr_config, book_out, sheet_roster, 0, sheet_name, filter, ROSTER_FIXUPS)

    # remove the default sheet in a new wb that we don't need
    del book_out['Sheet']

    # move roster to beginning of workbook
    # we should be using workbook.move_sheet(), but that didn't seem to work
    book_out.remove(sheet_roster)
    book_out._add_sheet(sheet_roster, index=0)

    # name of saved file
    roster_file_name = f"DR{ dr_config.dr_id } Staffing Report { report_date_stamp }.xlsx"
    roster_sps_file_name = f"DR{ dr_config.dr_id } SPS Report { report_date_stamp }.xlsx"

    # now figure out who we should send to
    roster_table_dicts = o365_config.read_table_to_dict('Roster', wb=book_out)
    mailing_list = o365_config.run_gap_patterns(roster_table_dicts)
    mailing_list_sps = o365_config.run_gap_patterns_sps(roster_table_dicts)

    # write out stuff to the config workbook
    last_report_date = o365_config.init_config_wb(roster_file_name, roster_sps_file_name, NOW, report_date)
    if last_report_date == report_date and args.ignore_too_soon != True:
        log.info(f"not running because report_date hasn't changed ({ report_date })")
        o365_config.update_report_status("Too Soon")
        return

    generate_contact_sheet(dr_config, roster_table_dicts, book_out, "Contacts")

    o365_config.init_recipient_sheet()

    o365_config.update_recipient_sheet(mailing_list, "Staffing")
    o365_config.update_recipient_sheet(mailing_list_sps, "SPS")

    if errors:
        sys.exit(1)

    # send to SPS
    do_distribution(dr_config, args, o365_config, book_out, "SPS Report", roster_sps_file_name,
                    report_date, mailing_list_sps)

    # send the regular roster out
    for sheet_name, filter in sps_sheets.items():
        # delete the sps sheet
        del book_out[sheet_name]

    do_distribution(dr_config, args, o365_config, book_out, "Staffing Report", roster_file_name,
                    report_date, mailing_list)

    o365_config.update_report_status("Success")



#
# generate a contact sheet that can be imported to google contacts
#
def generate_contact_sheet(dr_config, roster_table_dicts, wb, contact_sheet_name):

    # put this before the arrival sheet
    arrival_sheet = wb['Arrival']
    ws = wb.create_sheet(contact_sheet_name, wb.index(arrival_sheet))
    ws.sheet_properties.tabColor = SHEET_COLOR

    # tiny helper function to set cell contents
    def set_cell_value(row, col, value):
        ws.cell(row=row, column=col +1).value = value


    title_row = [ 'Name Prefix', 'First Name', 'Middle Name', 'Last Name', 'Name Suffix',
                 'Phonetic First Name', 'Phonetic Middle Name', 'Phonetic Last Name', 'Nickname',
                 'File As', 'E-mail 1 - Label', 'E-mail 1 - Value', 'Phone 1 - Label', 'Phone 1 - Value',
                 'Address 1 - Label', 'Address 1 - Country', 'Address 1 - Street', 'Address 1 - Extended Address',
                 'Address 1 - City', 'Address 1 - Region', 'Address 1 - Postal Code',
                 'Address 1 - PO Box', 'Organization Name', 'Organization Title', 'Organization Department',
                 'Birthday', 'Event 1 - Label', 'Event 1 - Value', 'Relation 1 - Label', 'Relation 1 - Value',
                 'Website 1 - Label', 'Website 1 - Value', 'Custom Field 1 - Label', 'Custom Field 1 - Value',
                 'Notes', 'Labels' ]

    for idx, value in enumerate(title_row):
        set_cell_value(1, idx, value)

    title_dict = spreadsheet_tools.title_to_dict(title_row)

    db_data = [ title_row ]
    row_num = 1

    for roster_dict in roster_table_dicts:
        row_num += 1
        name = roster_dict['Name']
        split = name.split(',')
        last_name = split[0].strip()
        first_name = split[1].strip()
        preferred_name = roster_dict['Preferred name']
        email = roster_dict['Email']
        phone = roster_dict['Cell phone']

        # don't bother adding if cell phone is blank
        if phone == '':
            continue

        # initialize the new row with the right number of elements
        #new_row = [ '' for x in range(len(title_row)) ]

        # set the data we care about
        set_cell_value(row_num, title_dict['First Name'], first_name)
        set_cell_value(row_num, title_dict['Last Name'], last_name)
        set_cell_value(row_num, title_dict['Nickname'], preferred_name)
        set_cell_value(row_num, title_dict['E-mail 1 - Label'], "Main")
        set_cell_value(row_num, title_dict['E-mail 1 - Value'], email)
        set_cell_value(row_num, title_dict['Phone 1 - Label'], "Mobile")
        set_cell_value(row_num, title_dict['Phone 1 - Value'], phone)
        set_cell_value(row_num, title_dict['Labels'], dr_config.dr_id)



#
# common code for sending both the SPS reports and the regular reports
#

def do_distribution(dr_config, args, o365_config, book_out, report_name, file_name, report_date, mailing_list):

    if args.send or args.test_send or args.save:
        log.debug(f"saving to { file_name }")
        book_out.save(file_name)
        o365_config.save_report_file(book_out, dr_config.reports_folder_name + "/" + file_name)

    try: 
        if args.send or args.test_send:
            send_roster(dr_config, args, o365_config.account, file_name, report_name, report_date, mailing_list)
    except:
        raise

    finally:
        if not args.save and (args.send or args.test_send):
            log.debug(f"removing { file_name }")
            os.remove(file_name)


#
# simple function that returns True if all elements of a list are None
#

def dt_convert(c):

    # definitely not a date
    if c == '' or c is None:
        return ''

    # already converted
    if isinstance(c, datetime.datetime):
        return c

    #new = spreadsheet_tools.excel_to_dt(c)
    new = xlrd.xldate_as_datetime(c, 0)

    #log.debug(f"dt_convert: orig '{ c }' new '{ new.strftime(REPORT_DATE_FORMAT) }'")
    return new


RIGHT_ALIGNED = openpyxl.styles.Alignment(horizontal="right")
ROSTER_FIXUPS = {
        'Name': { 'width': 25, },
        'Preferred name': { 'width': 10 },
        'Region': { 'width': 6, },
        'State': { 'width': 4, },
        'Res': { 'width': 4, },
        'T&M': { 'width': 6, },
        'GAP(s)': { 'width': 15, },
        'G/A/P': { 'width': 15, },
        'District': { 'width': 5, },
        'Qualification (assignment)': { 'width': 5, },
        'Current/Last Supervisor': { 'width': 30, },
        'Reporting/Work Location': { 'width': 30, },
        'On Job': { 'width': 5, },
        'DaysRemain': { 'width': 5, },
        '# dep': { 'width': 5, },
        'Lodging Last Night': { 'width': 30, },
        'Lodging Tonight': { 'width': 30, },
        'Qualifications (member)': { 'width': 30, },
        'All GAPs': { 'width': 30, },
        'All Supervisors': { 'width': 30, },
        'Work Location': { 'width': 30, },
        'Email': { 'width': 30, },
        'Assigned': { 'convert_value': dt_convert, 'number_format': "yyyy-mm-dd", },
        'Checked in': { 'convert_value': dt_convert, 'number_format': "yyyy-mm-dd", },
        'Released': { 'convert_value': dt_convert, 'number_format': "yyyy-mm-dd", },
        'Travel home': { 'convert_value': dt_convert, 'number_format': "yyyy-mm-dd", },
        'Last Daily Checkin': { 'convert_value': dt_convert, 'number_format': "yyyy-mm-dd", },
        'DaysRemain': { 'width': 5, 'number_format': "##0",
                       'alignment': RIGHT_ALIGNED,
                       #'convert_value': lambda x: x if x is not None and isinstance(x, int)) else x if  x == '' or x == 'n/a' else int(x),
                       'convert_value': lambda x: None if x is None else x if isinstance(x, int) else x if  x == '' or x == 'n/a' else int(x),
                       },
        'On Job': { 'width': 4, 'convert_value': lambda x: int(x) if isinstance(x, str) else x },
    }

ARRIVAL_FIXUPS = {
        'Region': { 'width': 6, },
        'Status': { 'width': 8, },
        'Resp': { 'width': 6, },
        'Category': { 'width': 6, },
        'Category': { 'width': 6, },
        'Gender': { 'width': 10, },
        'T&M': { 'width': 8, },
        'Trans': { 'width': 8, },
        'Flight Arrival Date/Time': { 'width': 8, },
        'Type': { 'width': 8, },
        'GAP': { 'width': 15, },
        '# Deploy': { 'width': 15, },
        'Texts?': { 'width': 8, },
        'Arrive date': { 'convert_value': dt_convert,
                     'number_format': "yyyy-mm-dd",
                     },
        'Flight Arrival Date/Time': { 'convert_value': dt_convert,
                     'number_format': "yyyy-mm-dd hh:mm",
                     'width': 18,
                     },
        }



AIR_FIXUPS = {
        'Departure City': { 'width': 16 },
        'Arrival City': { 'width': 16 },
        'Ticketed': { 'width': 4 },
        'Airline': { 'width': 15 },
        'Flight': { 'width': 8 },
        'Region name': { 'width': 24 },
        'Reporting or Work location': { 'width': 24 },
        'District': { 'width': 10 },

        'Last action date': { 'convert_value': dt_convert,
                     'number_format': "yyyy-mm-dd hh:mm",
                     'width': 18,
                     },
        'Exp Arrival': { 'convert_value': dt_convert,
                     'number_format': "yyyy-mm-dd",
                     },
        'Departure time': { 'convert_value': dt_convert,
                     'number_format': "yyyy-mm-dd hh:mm",
                     'width': 18,
                     },
        'Arrival time': { 'convert_value': dt_convert,
                     'number_format': "yyyy-mm-dd hh:mm",
                     'width': 18,
                     },
        }


SHIFTS_FIXUPS = {
        'Shift Name': { 'width': 30 },
        'County': { 'width': 24 },
        'Registration Status': { 'width': 16 },
        'Current Volunteer Status': { 'width': 20 },
        'District (of shift)': { 'width': 20 },
        'County (residence)': { 'width': 24 },
        'Registration Comments': { 'width': 24 },
        'Name': { 'width': 24 },
        'Registration Status': { 'width': 24 },
        'Ever DEBV/P-DEBV for this DRO': { 'width': 16 },
        'Email': { 'width': 24 },

        'Start Date': { 'convert_value': dt_convert,
                     'number_format': "yyyy-mm-dd",
                     },
        'Start Time': { 'convert_value': dt_convert,
                     'number_format': "yyyy-mm-dd hh:mm",
                     'width': 18,
                     },
        'End Date': { 'convert_value': dt_convert,
                     'number_format': "yyyy-mm-dd hh:mm",
                     'width': 18,
                     },
        'Date Registered/Last Changed': { 'convert_value': dt_convert,
                     'number_format': "yyyy-mm-dd hh:mm",
                     'width': 18,
                     },
        'End Time': { 'convert_value': dt_convert,
                     'number_format': "yyyy-mm-dd hh:mm",
                     'width': 18,
                     },
        }




def row_fixups(fixup_defs, label_row, supress_columns):


    fixups_by_col = {}

    # generate a mapping from column name to column index
    column_name_map = {}
    output_column = 0

    for c in range(0, len(label_row)):
        name = label_row[c]

        col_letter = openpyxl.utils.get_column_letter(c +1)
        if col_letter in supress_columns:
            # skip this column
            #log.debug(f"row_fixups: suppressing columm '{ col_letter }' c { c } output_column { output_column }")
            continue


        if name in fixup_defs:
            fixups_by_col[output_column] = fixup_defs[name]
        else:
            fixups_by_col[output_column] = {}

        column_name_map[name] = output_column
        output_column += 1


    return fixups_by_col, column_name_map




def fixup_cell_header(ws, c, fixup):

    col_letter = openpyxl.utils.get_column_letter(c +1)

    if 'width' in fixup:
        #log.debug(f"setting col { c } to { fixup['width'] }")
        ws.column_dimensions[col_letter].width = fixup['width']
    else:
        ws.column_dimensions[col_letter].auto_size = True


def fixup_cell(cell, fixup):

    if 'convert_value' in fixup:
        #log.debug(f"cell { cell } old value { cell.value } isint { isinstance(cell.value, int) }")
        cell.value = fixup['convert_value'](cell.value)
    if 'number_format' in fixup:
        cell.number_format = fixup['number_format']
    if 'alignment' in fixup:
        #log.debug(f"setting cell { cell } alignment { fixup['alignment'] }")
        cell.alignment = fixup['alignment']

# see if the responder is missing daily checkins.
# the rule: last checkin over 3 days, or if never checked in: DRO checkin date over 3 days
def filter_row_checkin_func(val, row, row_name_map, dr_config):

    checkin_warn_threshold = dr_config.late_checkin_threshold
    dt_threshold = NOW_NO_TZ - datetime.timedelta(days=checkin_warn_threshold)
    if isinstance(val, datetime.datetime):
        if val < dt_threshold:
            return True
    else:
        c = row_name_map['Checked in']
        val = row[c]
        if isinstance(val, datetime.datetime):
            if val < dt_threshold:
                return True

    return False


filter_row_active = { 'Released': lambda x: x == '' }
filter_row_overstayed = { 'DaysRemain': lambda x: x != '' and x != 'n/a' and x is not None and int(x) <= 0 }
filter_row_2_days = { 'DaysRemain': lambda x: x != '' and x != 'n/a' and x is not None and int(x) == 2 }
filter_row_needs_sup = { 'Current/Last Supervisor': lambda x: x != '' and x is not None and x == 'Needs Supervisor' }
filter_row_sms = { 'Texts?': lambda x: x != 'opt-in' }
filter_row_mc = { 'GAP(s)': lambda x: isinstance(x, str) and x.startswith('MC/') }
filter_row_cc = { 'GAP(s)': lambda x: isinstance(x, str) and x.startswith('CC/') }
filter_row_log = { 'GAP(s)': lambda x: isinstance(x, str) and x.startswith('LOG/') }
filter_row_er = { 'GAP(s)': lambda x: isinstance(x, str) and x.startswith('ER/') }
filter_row_ip = { 'GAP(s)': lambda x: isinstance(x, str) and x.startswith('IP/') }
filter_row_wf = { 'GAP(s)': lambda x: isinstance(x, str) and x.startswith('WF/') }
filter_row_om = { 'GAP(s)': lambda x: isinstance(x, str) and x.startswith('OM/') }

filter_row_checkin = { 'Last Daily Checkin': filter_row_checkin_func, 'extra_args': True }


def filter_row(row, row_name_map, filter_defs, dr_config):
    # check all the conditions in the filter; if they all pass include the row

    for name, func in filter_defs.items():
        if name == 'extra_args':
            # pseudo argument, not a column name.  Ignore
            continue

        #log.debug(f"filter_row: name '{ name }'")

        # ignore filtered columns
        if name in row_name_map:
            c = row_name_map[name]
            val = row[c]
            #log.debug(f"filter_row: name { name } c { c } val '{ val }' ")
            if 'extra_args' in filter_defs:
                pass
                if not func(val, row, row_name_map, dr_config):
                    #log.debug(f"filter_row: name { name } val '{ val }' returning False")
                    return False
            else:
                if not func(val):
                    #log.debug(f"filter_row: name { name } val '{ val }' returning False")
                    return False
        
    #log.debug(f"filter_row: returning True")
    return True


def read_roster(book_out, sheet_name, file_contents: str, label_row: int, fixups: dict, freeze_col: str = "B", suppress_columns: dict[str] = {}) -> openpyxl.worksheet.worksheet:
    
    book_in = xlrd.open_workbook(file_contents=file_contents)
    sheet_in = book_in.sheet_by_index(0)

    #log.debug(f"sheet name { sheet_in.name } rows { sheet_in.nrows } cols { sheet_in.ncols } label_row { label_row }")

    #for col in range(0, sheet.ncols):
    #    cell_value = sheet.cell_value(label_row, col)
    #    log.debug(f"cell({ label_row }, { col }) = { cell_value }")

    label_values = sheet_in.row_values(label_row)
    #log.debug(f"label_values: { label_values }")


    # copy everything to a clean workbook
    sheet_orig = book_out.create_sheet(sheet_name, 0)
    sheet_orig.title = sheet_name

    # set column attributes
    fixups_by_col, column_name_map = row_fixups(fixups, label_values, suppress_columns)
    output_col = 0
    for c in range(0, len(label_values)):
        col_letter = openpyxl.utils.get_column_letter(c +1)
        if col_letter in suppress_columns:
            continue

        if output_col not in fixups_by_col:
            log.error(f"read_roster: missing key: c { c } col_letter '{ col_letter }' output_col { output_col }")
        fixup_cell_header(sheet_orig, output_col, fixups_by_col[output_col])
        output_col += 1
        

    # copy cells
    for r in range(0, sheet_in.nrows):
        row = sheet_in.row_values(r)


        output_c = 0
        for c in range(0, sheet_in.ncols):
            col_letter = openpyxl.utils.get_column_letter(c +1)
            if col_letter in suppress_columns:
                # skip this column
                #log.debug(f"read_roster: suppressing columm '{ col_letter }'")
                continue

            
            value = row[c]
            cell = sheet_orig.cell(row=r +1, column=output_c +1, value=value)

            # don't fix up cells before the actual data
            if r > label_row:
                fixup_cell(cell, fixups_by_col[output_c])

            output_c = output_c + 1


    # make a table if there is data
    #log.debug(f"about to add table: sheet { sheet_name } sheet_orig.max_row { sheet_orig.max_row } label_row { label_row }")
    if sheet_orig.max_row > label_row:
        last_col_letter = openpyxl.utils.get_column_letter(sheet_orig.max_column)
        table_ref = f"A{label_row + 1}:{ last_col_letter }{ sheet_orig.max_row }"
        log.debug(f"adding table: table '{ sheet_name }' table_ref '{ table_ref }'")
        table = openpyxl.worksheet.table.Table(displayName=sheet_name, ref=table_ref)
        sheet_orig.add_table(table)

        sheet_orig.freeze_panes = f"{ freeze_col}{ label_row + 2 }"

    return sheet_orig


# copy from the 'orig' sheet to a new sheet, filtering entries
def copy_sheet(dr_config, wb, sheet_orig, label_row, sheet_name, filters, fixups,
               sheet_color: str = None,
               suppress_columns: dict[str] = {}):
    
    #log.debug(f"copy_sheet: sheet_name { sheet_name } label_row { label_row }")
    #sheet_new = wb.create_sheet(sheet_name, len(wb.sheetnames)-1)
    sheet_new = wb.create_sheet(sheet_name, 0)
    if sheet_color is not None:
        sheet_new.sheet_properties.tabColor = sheet_color

    label_values = list(next(sheet_orig.iter_rows(min_row=label_row +1, max_row=label_row +2, values_only=True)))

    # set column attributes
    fixups_by_col, column_name_map = row_fixups(fixups, label_values, suppress_columns)
    output_col = 0
    for c in range(0, len(label_values)):
        col_letter = openpyxl.utils.get_column_letter(c +1)
        if col_letter in suppress_columns:
            continue

        if output_col not in fixups_by_col:
            log.error(f"copy_sheet: c not in fixups_by col.  c { c } output_col { output_col } col_letter { col_letter }")
        fixup_cell_header(sheet_new, output_col, fixups_by_col[output_col])
        output_col += 1

    # these two are origin one indexes
    max_col = sheet_orig.max_column
    max_row = sheet_orig.max_row

    # copy cells
    output_row = 1
    for r in range(label_row, max_row):
        row = list(sheet_orig.iter_rows(min_row=r +1, max_row=r +2, values_only=True))
        row_values = list(row[0])
        #log.debug(f"copy_sheet: row { row }")

        include_row = False
        if r > label_row:
            include_row = filter_row(row_values, column_name_map, filters, dr_config)
        else:
            include_row = True

        if not include_row:
            #log.debug(f"row { r } output_row { output_row } not included")
            continue

        output_col = 0
        for c in range(0, max_col):
            col_letter = openpyxl.utils.get_column_letter(c+1)
            if col_letter in suppress_columns:
                # skip this column
                #log.debug(f"copy_sheet: suppressing columm '{ col_letter }'")
                continue

            cell = sheet_new.cell(row=output_row, column=output_col +1, value=row_values[c])

            # don't fix up cells before the actual data
            if r > label_row:
                fixup_cell(cell, fixups_by_col[output_col])
            output_col += 1

        # increment the output row
        output_row = output_row + 1


    # make a table if there is data
    if sheet_new.max_row > 2:
        last_col_letter = openpyxl.utils.get_column_letter(sheet_new.max_column)

        table_ref = f"A1:{ last_col_letter }{ sheet_new.max_row }"
        log.debug(f"copy_sheet: adding table { sheet_name } table_ref '{ table_ref }'")
        table_new = openpyxl.worksheet.table.Table(displayName=sheet_name, ref=table_ref)
        sheet_new.add_table(table_new)
        sheet_new.freeze_panes = f"B2"

    return sheet_new


#
# wrapper for sending out the roster
#
def send_roster(dr_config, args, account, file_name, report_type, report_date, mailing_list):

    warn_days = 2
    if report_date < NOW - datetime.timedelta(days=warn_days):
        date_warning = textwrap.dedent(
                f"""
                <p>
                <span style='background-color:yellow;'>WARNING: staff report is more than { warn_days } days old<span>.
                </p>
                """)
    else:
        date_warning = ""

    if report_type == "SPS Report":
        sps_section = textwrap.dedent(
                f"""
                <p>
                This is the SPS version of this report
                which several additional worksheets to slightly ease the SPS workflow.
                </p>

                <div style="padding-left:20px;">

                    <p>
                    The Outprocess worksheet has people with zero days (or negative days) of remaining time on the DR.
                    They should be contacted to outprocess or extend their deployment
                    </p>

                    <p>
                    The Days_2 worksheet has people with 2 days left on their deployment.
                    It is common to reach out to these folks with a template with outprocessing instructions
                    </p>

                    <p>
                    The Needs_Sup sheet shows people with their supervisor set to Needs Supervisor
                    </p>

                    <p>
                    The Need_SMS sheet shows people who have not opted in to DCS Text Messaging;
                    they should be contacted to opt in so they can receive DRIS and other text messages.
                    </p>

                    <p>
                    Late_checkin show people who haven't done a daily checkin in more than
                    { dr_config.late_checkin_threshold } days
                    </p>

                    <p>
                    The Contacts sheet is a bit different.  This can be exported as a CSV file and then
                    imported into Google contacts.  This will give phone number to name mapping for the
                    Google ecosystem.
                    </p>
                </div>

                """)
    else:
        sps_section = ""

    message_body = textwrap.dedent(
        f"""
        <p>
        This report is based on the staff roster dated { report_date.strftime(REPORT_DATE_FORMAT) }.
        </p>

        { date_warning }

        <p>
        Hello everyone.  This is a spreadsheet based on the automated staffing reports, but reorganized to hopefully be easier to use.
        </p>


        <p>
        It has the same data as the normal automated report, but has been reformatted for easier use.
        All the sheets have been made with tables, so the columns are sortable and filterable.
        The sheets have been 'frozen', so column headers and names are always visible,
        even when scrolling in the worksheet.
        </p>

        <p>
        There is a worksheet called "Roster" which has all responders who have not been out-processed appear.
        By default this is sorted by GAP, but you can easily re-sort it using the column headers
        </p>

        <p>
        In addition there is a worksheet for every Group, with the responders in that group in that worksheet.
        </p>

        <p>
        Finally there is a sheet for all the other automated reports,
        as well as the original version of the staff roster (called "Orig") that has outprocessed people also.
        </p>

        { sps_section }

        <p>
        DR { dr_config.dr_id } Staff Planning and Support
        </p>

        <p>
        If you have any questions or suggestions about the report or its contents, feel free to contact
        <a href='mailto:dr-report-automation@redcross.org'>dr-report-automation@redcross.org</a>.
        </p>

        """)

    send_report_common(dr_config, args, account, file_name, report_type, message_body, mailing_list)




def send_report_common(dr_config, args, account, file_name, report_type, message_body, mailing_list):

    if args.test_send:
        send_report_common2(dr_config, args, account, file_name, report_type, message_body,
                            dr_config.to_test, None, None)

    if args.send:
        for e in mailing_list:
            if isinstance(e, str):
                email = e
                gap = None
                name = None
            else:
                email = e['Email']
                gap = e['GAP(s)']
                name = e['Name']
            send_report_common2(dr_config, args, account, file_name, report_type, message_body, email, gap, name)



def send_report_common2(dr_config, args, account, file_name, report_type, message_body, email, gap, name):

    message = account.new_message(resource=dr_config.send_email)
    #message = account.new_message()

    #message.to.add(email)
    message.to.add(dr_config.to_test)

    #if extra_recipients != None and len(extra_recipients) > 0:
    #    log.debug(f"adding extra recipients { extra_recipients }")

    name_string = f"{ name } " if name is not None else ""
    gap_string = f"with GAP { gap }" if gap is not None else ""
    posting = textwrap.dedent(
            f"""
            <p>
            This message is being sent to { name_string }{ email }{ gap_string}.
            Please contact <a href='{ dr_config.send_email }'>{ dr_config.send_email }</a>
            if you do not want to receive future mailings
            """)

    if args.send:
        log.debug(f"sending { file_name } to { email }")
        debug_message = ""
    else:
        log.debug(f"debug sending { file_name } to { email }")
        debug_message = textwrap.dedent(
            f"""
            <p>
            DEBUG Version: not sent to the list
            </p>
            """)

    message.body = textwrap.dedent(
        f"""
        <!DOCTYPE html>
        <html>
        <meta http-equiv="Content-type" content="text/html" charset="UTF8" />
        <title>DR{ dr_config.dr_id } { report_type }</title>
        </head>
        <body>

        <h1>DR{ dr_config.dr_id } { report_type }</h1>
        { debug_message }
        { posting }

        { message_body }

        </body>
        </html>
        """)


    message.subject = file_name
    message.attachments.add( file_name )

    try:
        if not args.suppress_email:
            message.send(save_to_sent_folder=True)
    except requests.RequestException as e:
        log.error(f"got an error: { e }, response json { e.response.json }")
        raise e


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
            description="tool to convert staffing reports into a more usefull form",
            allow_abbrev=False)
    parser.add_argument("--debug", help="turn on debugging output", action="store_true")
    parser.add_argument("--suppress-email", help="don't actually send the email", action="store_true")
    parser.add_argument("--ignore-too-soon", help="run even if there is no new staff report", action="store_true")
    parser.add_argument("--save", help="retain output file", action="store_true")
    parser.add_argument("--send", help="send emails out", action="store_true")
    parser.add_argument("--test-send", help="send emails out, but to the test email box", action="store_true")
    parser.add_argument("--dr-id", help="Identifier for the DRO; must match the staffing report", required=True, action="store")

    args = parser.parse_args()

    return args


if __name__ == "__main__":
    neil_tools.init_logging(__name__)
    log = logging.getLogger(__name__)
    main()
else:
    log = logging.getLogger(__name__)
